import numpy as np
import pandas as pd
import openpyxl
import sklearn

file = '5ML-Riduzione-PCA-EsercizioGuidato-EsercizioAutonomia-Dati.xlsx'

# data = pd.read_excel(file, sheet_name=0, skiprows=2, usecols=(3,4,5,6,7,8,9), na_filter=False)
data = pd.read_excel(file, sheet_name=1, skiprows=2, usecols=(3,4,5,6,7,8), na_filter=False)
data.dropna(how='all', inplace=True)
print("PATTERN")
print(data)
print("\n")

#wb = openpyxl.load_workbook(file, data_only=True)
#sheets = wb.sheetnames
#page = 1
#sheet = wb[sheets[page]]
#print(sheets[page])

#data = []
#for row in sheet.iter_rows(values_only=True):
#    first_not_none = -1
#    last_not_none = -1
#    for i, cell in enumerate(row):
#        if cell is not None:
#            if first_not_none == -1:
#                first_not_none = i
#            last_not_none = i

#    if first_not_none != -1:
#        data.append(row[first_not_none:last_not_none+1])



#for d in data:
#    print(d)

cov = np.corrcoef(data, rowvar=False)
print("COVARIANCE")
print(cov)
print("\n")

eval, evec = np.linalg.eig(cov)
# evec = np.transpose(evec)

print("BEFORE SORTING")
print(eval)
print("\t")
print(evec)
print("\n")

sortindex = eval.argsort()[::-1]

eval = eval[sortindex]
evec = evec[:, sortindex]

print("AFTER SORTING")
print(eval)
print("\t")
print(evec)
print("\n")

tot = np.sum(eval)
sum = 0
perc = []
cumul = []

for i in eval:
    sum += i
    perc.append(i/tot)
    cumul.append(sum/tot)


print("Num\tEigen\tWeigth\tCumulative\t")
for i, e in enumerate(eval):
    #print(i, '\t', e, '\t', perc[i], '%\t', cumul[i], '%')
    print('{:} \t{:.2} \t{:.2%} \t{:.2%} '.format(i+1, e, perc[i], cumul[i]))
print("\n")

trans = evec[(0, 1, 2), :]
print(trans)