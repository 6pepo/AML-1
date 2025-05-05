import numpy as np
import pandas as pd
import openpyxl
import matplotlib.pyplot as plt

from scipy.linalg import eig

file = '5ML-Riduzione-PCA-EsercizioGuidato-EsercizioAutonomia-Dati.xlsx'

# data = pd.read_excel(file, sheet_name=0, skiprows=2, usecols=(3,4,5,6,7,8,9), na_filter=False)
# data.dropna(how='all', inplace=True)
# print(data)

wb = openpyxl.load_workbook(file, data_only=True)
sheets = wb.sheetnames
page = 1
sheet = wb[sheets[page]]
print(sheets[page])

data = []
for row in sheet.iter_rows(values_only=True):
    first_not_none = -1
    last_not_none = -1
    for i, cell in enumerate(row):
        if cell is not None:
            if first_not_none == -1:
                first_not_none = i
            last_not_none = i

    if first_not_none != -1:
        data.append(row[first_not_none:last_not_none+1])

data = np.array(data)
print('Pattern')
print(data,'\n')

corr = np.corrcoef(data, rowvar=False)
print('Correlation Matrix')
print(corr,'\n')

e_val, e_vec = np.linalg.eig(corr)
# e_val, e_vec = eig(corr) #in teoria più efficiente con scipy
e_vec = np.transpose(e_vec)

sort_index = np.argsort(e_val)[::-1]
e_val = e_val[sort_index]
e_vec = e_vec[sort_index,:]

for i in range(len(e_val)):
    print('Eigenvalue:', e_val[i])
    print('Eigenvector:', e_vec[i],'\n')

e_val_sum = np.sum(e_val)

val_sum = 0
print('\nPercent\tCumulative')
for i,val in enumerate(e_val):
    val_sum += val
    print(round(val/e_val_sum * 100, 2), '%\t', round(val_sum/e_val_sum * 100, 2), '%')

fig,ax = plt.subplots()
plt.plot(e_val, marker='o', color = 'red')
plt.show()

pc = 3
names = ['polla','chepo','echch','amare','xanst','polav']
print(f'\nFirst {pc} principal components')
for i in range(pc):
    line = ''
    for j in range(len(e_vec[i])):
       line += f' {round(e_vec[i,j],4)}*{names[j]} '    #{data[i,j]} '
    print(f'y_{i}: {line}')